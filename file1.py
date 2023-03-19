#!/usr/bin/env python
# coding: utf-8

# In[69]:


import tabula, pandas as pd, os, openpyxl as xl
import time
import datetime
import shutil
import getpass
import numpy as np
from openpyxl import Workbook
path = os.getcwd()
name_1 = str('ОДДС')
file = os.listdir(path)
wb2 = Workbook()
#print(file)
for f in file:
    #print(f)
    if (f.endswith('.xlsx') ) and (name_1 in f):
        #print(f)
        wb1 = xl.load_workbook(path + '\\' + f)
        ws1 = wb1.worksheets[0]
        #wb2 = Workbook()
        ws2 = wb2.create_sheet(ws1.title)

        for row in ws1:
            for cell in row:
                ws2[cell.coordinate].value = cell.value

        wb2.save('КПК_ОДДС.xlsx')


        
        
sheet_delete1 = wb2['Sheet']    
wb2.remove(sheet_delete1)
wb2.save('КПК_ОДДС.xlsx')
#sheet_delete2 = wb2['Sheet1']    
#wb2.remove(sheet_delete2)
#wb2.save('КПК_ОДДС.xlsx')
for sheet in wb2:
    print(sheet.title)
    for col in sheet.iter_cols():
        for cell in col:
            if cell.value == 'огрн':
                col == 5
            if cell.value == 'месяц':
                col == 1
            if cell.value == 'год':
                col == 2
    
            
            
wb2.save('КПК_ОДДС.xlsx')        
#x+=1        
    


# In[ ]:





# In[ ]:





# In[ ]:




